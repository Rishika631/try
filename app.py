import streamlit as st
from PIL import Image
import openpyxl
from openpyxl import Workbook
import requests
import json
import barcode
import numpy as np
import base64
from io import BytesIO


API_TOKEN = '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'  # Replace with your Hugging Face OCR API token

def extract_data(image):
    url = "https://api-inference.huggingface.co/models/huggingface/t5-base-ocr"
    headers = {
        'Authorization': f'Bearer {API_TOKEN}',
        'Content-Type': 'application/json'
    }

    # Convert the image to bytes
    image_byte_array = BytesIO()
    image.save(image_byte_array, format='PNG')
    image_byte_array = image_byte_array.getvalue()

    # Convert the image bytes to base64
    image_base64 = base64.b64encode(image_byte_array).decode('utf-8')

    # Prepare the API request payload
    payload = {
        "inputs": {
            "image": {
                "base64": image_base64
            }
        }
    }

    # Send the API request
    response = requests.post(url, headers=headers, data=json.dumps(payload))
    response_json = response.json()

    # Extract the OCR results
    extracted_data = []
    if 'predictions' in response_json:
        for result in response_json['predictions']:
            item = result['label']
            price = result['score']
            extracted_data.append((item, price))

    return extracted_data


def generate_barcode(item):
    # Generate a barcode image for the given item using pyBarcode
    barcode_image_path = f'{item}.png'
    barcode_image = barcode.get('code39', str(item), writer=ImageWriter()).render(writer_options={'module_width': 0.2, 'module_height': 15})
    with open(barcode_image_path, 'wb') as f:
        f.write(barcode_image)
    return barcode_image_path


def save_to_excel(data):
    wb = Workbook()
    sheet = wb.active

    # Write the data to the Excel sheet
    sheet['A1'] = 'Item'
    sheet['B1'] = 'Price'

    for row, (item, price) in enumerate(data, start=2):
        sheet.cell(row=row, column=1).value = item
        sheet.cell(row=row, column=2).value = price

    # Save the Excel file
    wb.save(filename='invoice_data.xlsx', file_format='xlsx')




def main():
    st.title("Invoice Processing App")
    st.write("Upload an invoice image and extract data")

    # Upload image
    uploaded_file = st.file_uploader("Choose an image", type=["jpg", "jpeg", "png"])

    if uploaded_file is not None:
        # Convert the uploaded image to PIL format
        image = Image.open(uploaded_file)

        # Display the uploaded image
        st.image(image, caption='Uploaded Image', use_column_width=True)

        # Extract data from the invoice
        extracted_data = extract_data(image)

        # Generate barcode for each item
        for item, price in extracted_data:
            barcode_image = generate_barcode(item)
            st.image(barcode_image, caption=f'Barcode for {item}', use_column_width=True)

        # Save data to Excel file
        save_to_excel(extracted_data)

        st.success("Data extraction and barcode generation completed.")
        st.download_button(
            label='Download Excel',
            data='invoice_data.xlsx',
            file_name='invoice_data.xlsx'
        )

if __name__ == '__main__':
    main()
